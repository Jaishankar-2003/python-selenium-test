from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.edge import service
from selenium.webdriver.edge.service import Service
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import requests as requests
import time
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import os

# driver = webdriver.Chrome()

# ---------------------------------------------------------------------------------------------------------------

# Data driven testing (excel , openpyxl .xlsx) ()

# read data to exl , write data to test , data driven test

file = "C:\\Users\\JAI-SHANKAR\\Downloads\\Orders-With Nulls.xlsx"
workbook=openpyxl.load_workbook(file)
print(workbook.sheetnames)
sheet=workbook['Sheet2']
print(sheet)

rows=sheet.max_row
col=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value,end='     ')
        print()